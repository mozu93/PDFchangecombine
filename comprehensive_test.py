#!/usr/bin/env python3
"""
PDF変換・結合ツール 包括的テストスクリプト
機能を段階的にテスト・デバッグを実行
"""

import sys
import os
import time
from pathlib import Path

# プロジェクトのsrcディレクトリをパスに追加
sys.path.insert(0, str(Path(__file__).parent / "src"))

from src.core.combiner import PDFCombiner
from src.core.converter import PDFConverter
from src.core.office_converter import OfficeConverter
from src.core.image_converter import ImageConverter
from src.utils.logger import logger
from src.utils.error_handler import error_handler, ErrorSeverity
from src.utils.file_utils import FileScanner

def print_separator(title: str):
    """区切り線表示"""
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")

def test_file_scanner():
    """ファイルスキャナーテスト"""
    print_separator("ファイルスキャナーテスト")
    
    try:
        # テストディレクトリの確認
        test_dir = Path("test_pdfs")
        if test_dir.exists():
            print(f"OK テストディレクトリ確認: {test_dir}")
            
            # ファイルスキャン実行
            paths = [str(test_dir)]
            result = FileScanner.scan_files_from_paths(paths)
            
            print(f"  - 有効ファイル: {len(result['valid'])}個")
            print(f"  - 無効ファイル: {len(result['invalid'])}個")
            print(f"  - 処理時間: {result['scan_time']:.3f}秒")
            
            for file_path in result['valid']:
                print(f"    OK {Path(file_path).name}")
            
            return True
        else:
            print("NG テストディレクトリが見つかりません")
            return False
            
    except Exception as e:
        print(f"NG ファイルスキャナーテストエラー: {e}")
        return False

def test_pdf_combination():
    """PDF結合機能テスト"""
    print_separator("PDF結合機能テスト")
    
    try:
        # テストファイルの準備
        test_dir = Path("test_pdfs")
        pdf_files = [
            str(test_dir / "test1.pdf"),
            str(test_dir / "test2.pdf"), 
            str(test_dir / "test3.pdf")
        ]
        
        output_path = str(test_dir / "combined_comprehensive.pdf")
        
        # 結合実行
        combiner = PDFCombiner()
        start_time = time.time()
        
        def progress_callback(message, progress):
            print(f"  進捗: {message} ({progress:.1f}%)")
        
        result = combiner.combine_pdfs(pdf_files, output_path, progress_callback)
        end_time = time.time()
        
        if result.success:
            print(f"OK 結合成功: {result.total_pages}ページ")
            print(f"  - 処理時間: {result.processing_time:.3f}秒")
            print(f"  - 出力ファイル: {output_path}")
            
            # 出力ファイル確認
            if Path(output_path).exists():
                size = Path(output_path).stat().st_size
                print(f"  - ファイルサイズ: {size} bytes")
            
            return True
        else:
            print(f"NG 結合失敗: {result.error_message}")
            return False
            
    except Exception as e:
        print(f"NG PDF結合テストエラー: {e}")
        return False

def test_office_conversion():
    """Office変換機能テスト"""
    print_separator("Office変換機能テスト") 
    
    try:
        converter = OfficeConverter()
        
        # テスト用Excelファイルの作成
        test_excel_path = Path("test_pdfs") / "test_excel.xlsx"
        
        # 簡単なExcelファイル作成（openpyxlがある場合）
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "テストシート"
            
            # データ入力
            ws['A1'] = "項目"
            ws['B1'] = "値"
            ws['A2'] = "テスト1"
            ws['B2'] = 100
            ws['A3'] = "テスト2"
            ws['B3'] = 200
            
            wb.save(test_excel_path)
            print(f"OK テスト用Excelファイル作成: {test_excel_path}")
            
            # PDF変換実行
            output_path = Path("test_pdfs") / "converted_excel.pdf"
            success = converter.convert_to_pdf(str(test_excel_path), str(output_path))
            
            if success and output_path.exists():
                print(f"OK Excel->PDF変換成功: {output_path}")
                print(f"  - ファイルサイズ: {output_path.stat().st_size} bytes")
                return True
            else:
                print("NG Excel->PDF変換失敗")
                return False
                
        except ImportError:
            print("NG openpyxlがインストールされていません")
            return False
            
    except Exception as e:
        print(f"NG Office変換テストエラー: {e}")
        return False

def test_image_conversion():
    """画像変換機能テスト"""
    print_separator("画像変換機能テスト")
    
    try:
        converter = ImageConverter()
        
        # テスト用画像の作成（Pillowがある場合）
        try:
            from PIL import Image, ImageDraw
            
            # 簡単なテスト画像作成
            test_image_path = Path("test_pdfs") / "test_image.png"
            
            # 200x200の白い背景に赤い四角形
            img = Image.new('RGB', (200, 200), 'white')
            draw = ImageDraw.Draw(img)
            draw.rectangle([50, 50, 150, 150], fill='red', outline='black')
            draw.text((75, 175), "TEST", fill='black')
            
            img.save(test_image_path)
            print(f"OK テスト用画像作成: {test_image_path}")
            
            # PDF変換実行
            output_path = Path("test_pdfs") / "converted_image.pdf"
            success = converter.convert_to_pdf(str(test_image_path), str(output_path))
            
            if success and output_path.exists():
                print(f"OK 画像->PDF変換成功: {output_path}")
                print(f"  - ファイルサイズ: {output_path.stat().st_size} bytes")
                return True
            else:
                print("NG 画像->PDF変換失敗")
                return False
                
        except ImportError:
            print("NG Pillowがインストールされていません")
            return False
            
    except Exception as e:
        print(f"NG 画像変換テストエラー: {e}")
        return False

def test_error_handling():
    """エラーハンドリングテスト"""
    print_separator("エラーハンドリングテスト")
    
    try:
        # 意図的にエラーを発生させてテスト
        test_error = FileNotFoundError("テスト用ファイルが見つかりません")
        
        print("OK WARNING レベルエラーテスト")
        error_handler.handle_error(test_error, ErrorSeverity.WARNING, "テスト", "これはテスト用の警告です")
        
        print("OK INFO レベルエラーテスト")
        error_handler.handle_error(test_error, ErrorSeverity.INFO, "テスト", "これはテスト用の情報です")
        
        # 統計情報確認
        stats = error_handler.get_error_statistics()
        print(f"OK エラー統計: エラー{stats['error_count']}回, 警告{stats['warning_count']}回")
        
        return True
        
    except Exception as e:
        print(f"NG エラーハンドリングテストエラー: {e}")
        return False

def test_performance():
    """パフォーマンステスト"""
    print_separator("パフォーマンステスト")
    
    try:
        # 起動時間測定
        start_time = time.time()
        
        # 各コンポーネントの初期化時間測定
        converter_start = time.time()
        converter = PDFConverter()
        converter_time = time.time() - converter_start
        
        combiner_start = time.time()
        combiner = PDFCombiner()
        combiner_time = time.time() - combiner_start
        
        total_time = time.time() - start_time
        
        print(f"OK PDFConverter初期化: {converter_time:.3f}秒")
        print(f"OK PDFCombiner初期化: {combiner_time:.3f}秒")
        print(f"OK 総初期化時間: {total_time:.3f}秒")
        
        # 要件定義書の最大起動時間（5秒）チェック
        from src.config import MAX_STARTUP_TIME_SECONDS
        if total_time <= MAX_STARTUP_TIME_SECONDS:
            print(f"OK 起動時間要件クリア (上限: {MAX_STARTUP_TIME_SECONDS}秒)")
            return True
        else:
            print(f"NG 起動時間要件違反 (上限: {MAX_STARTUP_TIME_SECONDS}秒, 実測: {total_time:.3f}秒)")
            return False
            
    except Exception as e:
        print(f"NG パフォーマンステストエラー: {e}")
        return False

def run_comprehensive_test():
    """包括的テスト実行"""
    print_separator("PDF変換・結合ツール 包括的テスト開始")
    
    tests = [
        ("ファイルスキャナー", test_file_scanner),
        ("PDF結合機能", test_pdf_combination),
        ("Office変換機能", test_office_conversion),
        ("画像変換機能", test_image_conversion),
        ("エラーハンドリング", test_error_handling),
        ("パフォーマンス", test_performance),
    ]
    
    results = {}
    total_tests = len(tests)
    passed_tests = 0
    
    for test_name, test_func in tests:
        print(f"\n[{passed_tests + 1}/{total_tests}] {test_name}テスト実行中...")
        
        try:
            success = test_func()
            results[test_name] = success
            if success:
                passed_tests += 1
                print(f"[OK] {test_name}テスト: 成功")
            else:
                print(f"[NG] {test_name}テスト: 失敗")
        except Exception as e:
            print(f"[NG] {test_name}テスト: 例外発生 - {e}")
            results[test_name] = False
    
    # 結果サマリー
    print_separator("テスト結果サマリー")
    print(f"総テスト数: {total_tests}")
    print(f"成功: {passed_tests}")
    print(f"失敗: {total_tests - passed_tests}")
    print(f"成功率: {(passed_tests/total_tests)*100:.1f}%")
    
    print("\n詳細結果:")
    for test_name, success in results.items():
        status = "[OK] 成功" if success else "[NG] 失敗"
        print(f"  {test_name}: {status}")
    
    # 総合判定
    if passed_tests == total_tests:
        print(f"\n[完了] 全テスト成功！アプリケーションは正常に動作しています。")
        return True
    else:
        print(f"\n[警告] {total_tests - passed_tests}個のテストが失敗しました。修正が必要です。")
        return False

if __name__ == "__main__":
    success = run_comprehensive_test()
    sys.exit(0 if success else 1)